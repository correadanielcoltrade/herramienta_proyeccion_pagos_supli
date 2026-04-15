import io
from datetime import datetime
from flask import Blueprint, jsonify, request, send_file
from blueprint.security import auth_required
from blueprint.storage import load_records, save_records, utc_now_iso
from blueprint.analytics import build_summary, build_week_summary, build_provider_summary, build_gantt_data_with_refs
from blueprint.cache import get_cached, set_cached, invalidate_cache
from blueprint.excel import EXCEL_MIME

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    openpyxl = None

dashboard_bp = Blueprint('dashboard', __name__)

def _ensure_openpyxl():
    if not openpyxl:
        raise RuntimeError('openpyxl no disponible')
    return openpyxl


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _normalize_category(value):
    if not value:
        return 'Normal'
    key = str(value).strip().upper()
    return {'ALTA': 'Alta', 'MEDIA': 'Media', 'BAJA': 'Baja', 'NORMAL': 'Normal'}.get(key, 'Normal')


def _priority_from_categories(provider_cat, product_cat):
    provider_norm = _normalize_category(provider_cat)
    product_norm = _normalize_category(product_cat)
    provider_level = provider_norm if provider_norm in ('Alta', 'Media', 'Baja') else 'Baja'
    product_level = product_norm if product_norm in ('Alta', 'Media', 'Baja') else 'Baja'

    if product_level == 'Alta':
        return 'Alta'
    if product_level == 'Media':
        return 'Alta' if provider_level == 'Alta' else 'Media'
    return 'Baja' if provider_level == 'Baja' else 'Media'


def _normalize_status(value):
    if not value:
        return 'Pendiente'
    key = str(value).strip().upper()
    return 'Pagado' if key == 'PAGADO' else 'Pendiente'

@dashboard_bp.route('/dashboard/summary', methods=['GET'])
@auth_required()
def dashboard_summary():
    cache_key = 'dashboard:summary'
    cached = get_cached(cache_key)
    if cached:
        return jsonify(cached), 200

    payments = load_records('pp_payments')
    summary = build_summary(payments)
    set_cached(cache_key, summary, ttl=30)
    return jsonify(summary), 200


@dashboard_bp.route('/dashboard/weeks', methods=['GET'])
@auth_required()
def dashboard_weeks():
    cache_key = 'dashboard:weeks'
    cached = get_cached(cache_key)
    if cached:
        return jsonify(cached), 200

    payments = load_records('pp_payments')
    weeks = build_week_summary(payments)
    set_cached(cache_key, weeks, ttl=30)
    return jsonify(weeks), 200


@dashboard_bp.route('/dashboard/providers', methods=['GET'])
@auth_required()
def dashboard_providers():
    cache_key = 'dashboard:providers'
    cached = get_cached(cache_key)
    if cached:
        return jsonify(cached), 200

    payments = load_records('pp_payments')
    provider_records = load_records('pp_providers')
    providers = build_provider_summary(payments, provider_records=provider_records)
    set_cached(cache_key, providers, ttl=30)
    return jsonify(providers), 200


@dashboard_bp.route('/dashboard/gantt', methods=['GET'])
@auth_required()
def dashboard_gantt():
    cache_key = 'dashboard:gantt'
    cached = get_cached(cache_key)
    if cached:
        return jsonify(cached), 200

    payments = load_records('pp_payments')
    providers = load_records('pp_providers')
    gantt = build_gantt_data_with_refs(payments, providers=providers)
    set_cached(cache_key, gantt, ttl=30)
    return jsonify(gantt), 200


@dashboard_bp.route('/dashboard/gantt/status', methods=['PATCH'])
@auth_required()
def dashboard_gantt_status():
    payload = request.get_json(silent=True) or {}
    provider_id = payload.get('provider_id')
    week = payload.get('week')
    week_year = payload.get('week_year')
    status = _normalize_status(payload.get('status'))

    if provider_id is None or week is None:
        return jsonify({'error': 'Proveedor y semana son requeridos'}), 400

    payments = load_records('pp_payments')
    updated = 0

    for payment in payments:
        data = payment.get('data_json', {})
        payment_week = data.get('week')
        payment_week_year = data.get('week_year')
        if payment_week is None:
            parsed = _parse_date(data.get('date'))
            if parsed:
                payment_week_year, payment_week, _ = parsed.isocalendar()
        if str(data.get('provider_id')) != str(provider_id):
            continue
        if str(payment_week) != str(week):
            continue
        if week_year is not None and payment_week_year is not None:
            if str(payment_week_year) != str(week_year):
                continue
        if data.get('status') == status:
            continue
        data['status'] = status
        payment['updated_at'] = utc_now_iso()
        updated += 1

    if updated:
        save_records('pp_payments', payments)
        invalidate_cache('dashboard')

    return jsonify({'updated': updated, 'status': status}), 200


@dashboard_bp.route('/dashboard/gantt/export', methods=['GET'])
@auth_required()
def dashboard_gantt_export():
    _ensure_openpyxl()
    week_filter = request.args.get('week')

    payments = load_records('pp_payments')
    providers = load_records('pp_providers')
    products = load_records('pp_products')

    provider_index = {str(p.get('id')): p for p in providers}
    product_index = {str(p.get('id')): p for p in products}

    rows = {}
    priority_rank = {'Baja': 1, 'Media': 2, 'Alta': 3}

    for payment in payments:
        data = payment.get('data_json', {})
        week = data.get('week')
        week_year = data.get('week_year')
        if not week:
            parsed = _parse_date(data.get('date'))
            if parsed:
                week_year, week, _ = parsed.isocalendar()

        if week_filter and str(week) != str(week_filter):
            continue

        provider_id = data.get('provider_id')
        provider = provider_index.get(str(provider_id)) if provider_id is not None else None
        provider_data = provider.get('data_json', {}) if provider else {}
        provider_name = (provider_data.get('name') if provider else None) or data.get('provider_name') or 'Sin proveedor'
        provider_category = _normalize_category(provider_data.get('category') or data.get('provider_category'))

        items = data.get('items') or []
        is_paid = data.get('status') == 'Pagado'
        if not items:
            invoices = data.get('invoices') if isinstance(data.get('invoices'), list) else []
            product_names = data.get('product_names') if isinstance(data.get('product_names'), list) else []
            items = [{
                'invoice': invoices[0] if invoices else None,
                'product_name': product_names[0] if product_names else data.get('product_name'),
                'product_category': data.get('product_category'),
                'amount': data.get('amount') or 0,
            }]

        for item in items:
            if not isinstance(item, dict):
                continue
            product_id = item.get('product_id')
            product = product_index.get(str(product_id)) if product_id is not None else None
            product_data = product.get('data_json', {}) if product else {}
            product_name = (product_data.get('label') or product_data.get('description') or product_data.get('name')) if product else item.get('product_name')
            if not product_name:
                product_name = 'Producto -'

            invoice = item.get('invoice') or 'Factura -'
            product_category = _normalize_category((product_data.get('category') if product else None) or item.get('product_category') or data.get('product_category'))

            amount = item.get('amount')
            if amount is None:
                amount = item.get('quantity')
            amount = float(amount or 0)
            if amount == 0 and len(items) == 1:
                amount = float(data.get('amount', 0) or 0)

            priority = _priority_from_categories(provider_category, product_category)

            if week:
                week_label = f"Semana {week} ({week_year})" if week_year else f"Semana {week}"
                sort_value = int(week_year or 0) * 100 + int(week)
            else:
                week_label = 'Semana -'
                sort_value = 999999

            key = (sort_value, week_label, provider_name, str(invoice))
            entry = rows.get(key)
            if not entry:
                rows[key] = {
                    'total': amount,
                    'paid_total': amount if is_paid else 0,
                    'pending_total': 0 if is_paid else amount,
                    'priority': priority,
                }
            else:
                entry['total'] += amount
                if is_paid:
                    entry['paid_total'] += amount
                else:
                    entry['pending_total'] += amount
                if priority_rank.get(priority, 0) > priority_rank.get(entry['priority'], 0):
                    entry['priority'] = priority

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Programacion pagos'

    headers = ['Semana', 'Proveedor', 'Factura', 'Pagado', 'Pendiente', 'Total', 'Prioridad']
    sheet.append(headers)

    header_fill = PatternFill(fill_type='solid', fgColor='F6F1E7')
    header_font = Font(bold=True, color='4B443B')
    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center')

    priority_fills = {
        'Alta': PatternFill(fill_type='solid', fgColor='F4CCCC'),
        'Media': PatternFill(fill_type='solid', fgColor='FFF2CC'),
        'Baja': PatternFill(fill_type='solid', fgColor='D9EAD3'),
    }

    row_index = 2
    for key in sorted(rows.keys(), key=lambda k: (k[0], k[2].lower(), k[3].lower())):
        _, week_label, provider_name, invoice = key
        entry = rows[key]
        sheet.append([
            week_label,
            provider_name,
            invoice,
            entry['paid_total'],
            entry['pending_total'],
            entry['total'],
            entry['priority'],
        ])

        paid_cell = sheet.cell(row=row_index, column=4)
        paid_cell.number_format = '#,##0.00'
        paid_cell.alignment = Alignment(horizontal='right')

        pending_cell = sheet.cell(row=row_index, column=5)
        pending_cell.number_format = '#,##0.00'
        pending_cell.alignment = Alignment(horizontal='right')

        total_cell = sheet.cell(row=row_index, column=6)
        total_cell.number_format = '#,##0.00'
        total_cell.alignment = Alignment(horizontal='right')

        priority_cell = sheet.cell(row=row_index, column=7)
        fill = priority_fills.get(entry['priority'])
        if fill:
            priority_cell.fill = fill
        priority_cell.alignment = Alignment(horizontal='center')
        row_index += 1

    sheet.freeze_panes = 'A2'

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in range(2, row_index):
            value = sheet.cell(row=row, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(45, max(12, max_len + 2))

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"reporte_programacion_semana_{week_filter}.xlsx" if week_filter else "reporte_programacion_todas.xlsx"
    return send_file(
        buffer,
        mimetype=EXCEL_MIME,
        as_attachment=True,
        download_name=filename
    )

