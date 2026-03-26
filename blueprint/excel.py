import io

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None


EXCEL_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _ensure_openpyxl():
    if not openpyxl:
        raise RuntimeError('openpyxl no disponible')
    return openpyxl


def load_excel_records(file_storage):
    _ensure_openpyxl()
    workbook = openpyxl.load_workbook(file_storage, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = []
    for cell in rows[0]:
        if cell is None:
            headers.append('')
        else:
            headers.append(str(cell).strip())
    data = []
    for row in rows[1:]:
        record = {}
        has_value = False
        for idx, value in enumerate(row):
            key = headers[idx] if idx < len(headers) and headers[idx] else f'col_{idx}'
            record[key] = value
            if value not in (None, ''):
                has_value = True
        if has_value:
            data.append(record)
    return data


def build_excel_template(headers, sample_rows=None, sheet_name='Template'):
    _ensure_openpyxl()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = (sheet_name or 'Template')[:31]
    sheet.append(headers)
    for row in sample_rows or []:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
